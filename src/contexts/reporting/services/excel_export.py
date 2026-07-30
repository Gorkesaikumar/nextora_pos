import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from django.http import HttpResponse

class ExcelExportService:
    @staticmethod
    def generate(title: str, headers: list[str], rows: list[list]):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = title[:31]  # Excel sheet title limit is 31 chars
        
        # Style Definitions
        header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        align_center = Alignment(horizontal="center", vertical="center")
        
        # Write Headers
        ws.append(headers)
        for col_num, cell in enumerate(ws[1], 1):
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align_center
            
        # Write Rows
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, cell_val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=cell_val)
                # Format numbers nicely
                if isinstance(cell_val, (int, float)):
                    cell.number_format = '#,##0.00'
                    
        # Auto-adjust column widths
        for col_num, col in enumerate(ws.columns, 1):
            max_len = 0
            col_letter = get_column_letter(col_num)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_len:
                        max_len = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_len + 2)
            ws.column_dimensions[col_letter].width = min(adjusted_width, 50)  # Max width 50
            
        # Prepare response
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{title}.xlsx"'
        return response
